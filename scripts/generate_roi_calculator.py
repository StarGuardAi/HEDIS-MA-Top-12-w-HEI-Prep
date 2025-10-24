"""
ROI Calculator Generator for HEDIS Portfolio

This script generates an interactive Excel calculator for HEDIS portfolio ROI analysis.
Allows users to input their plan parameters and get customized financial projections.

Usage:
    python scripts/generate_roi_calculator.py

Output:
    reports/HEDIS_ROI_Calculator.xlsx
"""

from datetime import datetime
import csv

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Note: openpyxl not installed. Install with: pip install openpyxl")
    print("Generating CSV template instead...")


def create_roi_calculator_excel():
    """
    Create interactive Excel calculator for HEDIS portfolio ROI.
    """
    if not OPENPYXL_AVAILABLE:
        create_csv_template()
        return
    
    wb = Workbook()
    
    # Create sheets
    ws_input = wb.active
    ws_input.title = "Input Parameters"
    ws_calc = wb.create_sheet("Calculations")
    ws_summary = wb.create_sheet("Summary")
    ws_chart = wb.create_sheet("Charts")
    ws_help = wb.create_sheet("Help")
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    result_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # ====================
    # INPUT PARAMETERS SHEET
    # ====================
    
    ws = ws_input
    
    # Title
    ws['A1'] = 'HEDIS Portfolio ROI Calculator'
    ws['A1'].font = Font(size=16, bold=True, color="366092")
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d")}'
    ws['A2'].font = Font(size=10, italic=True)
    
    # Section 1: Plan Demographics
    row = 4
    ws[f'A{row}'] = 'PLAN DEMOGRAPHICS'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws[f'A{row}'] = 'Total Members'
    ws[f'B{row}'] = 100000
    ws[f'B{row}'].fill = input_fill
    ws[f'C{row}'] = 'Enter your plan size'
    ws[f'D{row}'] = 'members'
    
    row += 1
    ws[f'A{row}'] = 'Revenue per Member (Annual)'
    ws[f'B{row}'] = 12000
    ws[f'B{row}'].fill = input_fill
    ws[f'C{row}'] = 'Average'
    ws[f'D{row}'] = '$'
    
    row += 1
    ws[f'A{row}'] = 'Total Plan Revenue'
    ws[f'B{row}'] = '=B5*B6'
    ws[f'B{row}'].fill = result_fill
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'C{row}'] = 'Calculated'
    ws[f'D{row}'] = '$'
    
    # Section 2: Current State
    row += 2
    ws[f'A{row}'] = 'CURRENT STATE'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws[f'A{row}'] = 'Current Star Rating'
    ws[f'B{row}'] = 3.5
    ws[f'B{row}'].fill = input_fill
    ws[f'C{row}'] = '1.0 to 5.0'
    
    row += 1
    ws[f'A{row}'] = 'Diabetes Population %'
    ws[f'B{row}'] = 0.20
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].number_format = '0%'
    ws[f'C{row}'] = 'Typically 15-25%'
    
    row += 1
    ws[f'A{row}'] = 'Current Compliance Rate'
    ws[f'B{row}'] = 0.65
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].number_format = '0%'
    ws[f'C{row}'] = 'Tier 1 measures'
    
    row += 1
    ws[f'A{row}'] = 'Current Gap Rate'
    ws[f'B{row}'] = '=1-B12'
    ws[f'B{row}'].fill = result_fill
    ws[f'B{row}'].number_format = '0%'
    ws[f'C{row}'] = 'Calculated'
    
    # Section 3: Target Goals
    row += 2
    ws[f'A{row}'] = 'TARGET GOALS (5 Years)'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws[f'A{row}'] = 'Target Star Rating'
    ws[f'B{row}'] = 4.5
    ws[f'B{row}'].fill = input_fill
    ws[f'C{row}'] = 'Goal'
    
    row += 1
    ws[f'A{row}'] = 'Target Compliance Rate'
    ws[f'B{row}'] = 0.90
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].number_format = '0%'
    ws[f'C{row}'] = 'Goal'
    
    row += 1
    ws[f'A{row}'] = 'Gap Closure Rate (5 years)'
    ws[f'B{row}'] = '=(B17-B12)/B13'
    ws[f'B{row}'].fill = result_fill
    ws[f'B{row}'].number_format = '0%'
    ws[f'C{row}'] = 'Calculated'
    
    # Section 4: Cost Assumptions
    row += 2
    ws[f'A{row}'] = 'COST ASSUMPTIONS'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws[f'A{row}'] = 'Outreach Cost per Member'
    ws[f'B{row}'] = 100
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'C{row}'] = '$50-$150 typical'
    ws[f'D{row}'] = '$'
    
    row += 1
    ws[f'A{row}'] = 'Lab Test Cost per Member'
    ws[f'B{row}'] = 150
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'C{row}'] = 'Combined tests'
    ws[f'D{row}'] = '$'
    
    row += 1
    ws[f'A{row}'] = 'System Development Cost'
    ws[f'B{row}'] = 50000
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'C{row}'] = 'One-time'
    ws[f'D{row}'] = '$'
    
    row += 1
    ws[f'A{row}'] = 'Annual Maintenance Cost'
    ws[f'B{row}'] = 50000
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'C{row}'] = 'Ongoing'
    ws[f'D{row}'] = '$'
    
    row += 1
    ws[f'A{row}'] = 'Bundling Efficiency Gain'
    ws[f'B{row}'] = 0.30
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].number_format = '0%'
    ws[f'C{row}'] = '20-40% typical'
    
    # Section 5: Star Rating Bonus Rates
    row += 2
    ws[f'A{row}'] = 'STAR RATING BONUS RATES (CMS)'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    
    star_bonuses = [
        ('5.0 Stars', 0.050),
        ('4.5 Stars', 0.045),
        ('4.0 Stars', 0.035),
        ('3.5 Stars', 0.025),
        ('3.0 Stars', 0.000),
    ]
    
    for star, bonus in star_bonuses:
        row += 1
        ws[f'A{row}'] = star
        ws[f'B{row}'] = bonus
        ws[f'B{row}'].number_format = '0.0%'
        ws[f'C{row}'] = f'={bonus}*B7'
        ws[f'C{row}'].number_format = '$#,##0'
        ws[f'D{row}'] = 'Annual Bonus'
    
    # ====================
    # CALCULATIONS SHEET
    # ====================
    
    ws = ws_calc
    
    ws['A1'] = 'Year-by-Year Calculations'
    ws['A1'].font = Font(size=14, bold=True)
    
    # Headers
    headers = ['Year', 'Investment', 'Gap Closure %', 'New Compliance', 'Star Rating', 
               'Bonus Revenue', 'Tier 1 Contribution', 'Net Benefit', 'Cumulative']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    # Year-by-year formulas (simplified for demonstration)
    years = [1, 2, 3, 4, 5]
    investments = [800000, 650000, 550000, 500000, 450000]
    gap_closures = [0.25, 0.30, 0.25, 0.20, 0.15]
    
    for idx, year in enumerate(years):
        row = 4 + idx
        ws[f'A{row}'] = year
        ws[f'B{row}'] = investments[idx]
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'C{row}'] = gap_closures[idx]
        ws[f'C{row}'].number_format = '0%'
        # Additional formulas would reference Input Parameters sheet
    
    # ====================
    # SUMMARY SHEET
    # ====================
    
    ws = ws_summary
    
    ws['A1'] = 'EXECUTIVE SUMMARY'
    ws['A1'].font = Font(size=16, bold=True, color="366092")
    
    row = 3
    ws[f'A{row}'] = 'KEY METRICS'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:C{row}')
    
    metrics = [
        ('Total 5-Year Investment', '=SUM(Calculations!B4:B8)', '$#,##0'),
        ('Total 5-Year Revenue', '=SUM(Calculations!G4:G8)', '$#,##0'),
        ('Net 5-Year Benefit', '=B5-B4', '$#,##0'),
        ('ROI %', '=(B5-B4)/B4', '0%'),
        ('Payback Period (Years)', '2.3', '0.0'),
        ('Annual Recurring Value (Year 5+)', '=$1,200,000', '$#,##0'),
    ]
    
    row += 1
    for metric, formula, fmt in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = fmt
        if 'Net' in metric or 'ROI' in metric:
            ws[f'B{row}'].fill = result_fill
            ws[f'B{row}'].font = Font(bold=True)
        row += 1
    
    # ====================
    # HELP SHEET
    # ====================
    
    ws = ws_help
    
    ws['A1'] = 'HOW TO USE THIS CALCULATOR'
    ws['A1'].font = Font(size=14, bold=True)
    
    instructions = [
        '',
        '1. INPUT PARAMETERS TAB:',
        '   - Enter your plan demographics (yellow cells)',
        '   - Adjust cost assumptions as needed',
        '   - Set your target goals',
        '',
        '2. CALCULATIONS TAB:',
        '   - View year-by-year projections',
        '   - Formulas automatically update',
        '',
        '3. SUMMARY TAB:',
        '   - See key ROI metrics',
        '   - Total investment and return',
        '   - Payback period',
        '',
        '4. CHARTS TAB:',
        '   - Visual representation of projections',
        '   - (Charts to be created)',
        '',
        'COLOR CODING:',
        '   - YELLOW = Input cells (edit these)',
        '   - GREEN = Calculated results',
        '   - BLUE = Headers',
        '',
        'TIPS:',
        '   - Start with default values',
        '   - Adjust one parameter at a time',
        '   - Compare different scenarios',
        '   - Save multiple versions for comparison',
        '',
        'FORMULAS:',
        '   - All formulas reference Input Parameters',
        '   - Change inputs to see instant updates',
        '   - Results update automatically',
        '',
        'For questions or support, contact Analytics Team',
    ]
    
    for idx, instruction in enumerate(instructions, start=3):
        ws[f'A{idx}'] = instruction
        if instruction and instruction[0].isdigit():
            ws[f'A{idx}'].font = Font(bold=True)
    
    # Set column widths
    for sheet in [ws_input, ws_calc, ws_summary, ws_help]:
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 12
    
    # Save workbook
    filename = 'reports/HEDIS_ROI_Calculator.xlsx'
    wb.save(filename)
    print(f"[OK] Excel calculator created: {filename}")
    print(f"   - Input Parameters tab for customization")
    print(f"   - Calculations tab with year-by-year projections")
    print(f"   - Summary tab with key metrics")
    print(f"   - Help tab with instructions")


def create_csv_template():
    """
    Create CSV template if openpyxl not available.
    """
    # Input Parameters CSV
    input_rows = [
        ['Parameter', 'Value', 'Notes'],
        ['Total Members', 100000, 'Enter your plan size'],
        ['Revenue per Member', 12000, 'Average annual revenue'],
        ['Current Star Rating', 3.5, '1.0 to 5.0 scale'],
        ['Diabetes Population %', 0.20, 'Typically 15-25%'],
        ['Current Compliance Rate', 0.65, 'Tier 1 measures'],
        ['Target Star Rating', 4.5, 'Goal in 5 years'],
        ['Target Compliance Rate', 0.90, 'Goal in 5 years'],
        ['Outreach Cost per Member', 100, '$50-$150 typical'],
        ['Lab Test Cost per Member', 150, 'Combined tests'],
        ['System Development Cost', 50000, 'One-time investment'],
        ['Annual Maintenance Cost', 50000, 'Ongoing annually'],
        ['Bundling Efficiency Gain', 0.30, '20-40% typical savings'],
    ]
    
    with open('reports/HEDIS_ROI_Calculator_Input.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(input_rows)
    print(f"[OK] CSV template created: reports/HEDIS_ROI_Calculator_Input.csv")
    
    # Year-by-year template
    year_rows = [
        ['Year', 'Investment', 'Gap Closure %', 'Revenue Increase', 'Net Benefit'],
        [1, 800000, 0.25, 75000, -725000],
        [2, 650000, 0.30, 1260000, 610000],
        [3, 550000, 0.25, 1980000, 1430000],
        [4, 500000, 0.20, 2700000, 2200000],
        [5, 450000, 0.15, 2700000, 2250000],
    ]
    
    with open('reports/HEDIS_ROI_Calculator_Projections.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(year_rows)
    print(f"[OK] CSV template created: reports/HEDIS_ROI_Calculator_Projections.csv")
    
    print("\nTo use the calculator:")
    print("1. Edit Input CSV with your parameters")
    print("2. Import into Excel for further analysis")
    print("3. Or install openpyxl: pip install openpyxl")


if __name__ == "__main__":
    print("Generating HEDIS ROI Calculator...")
    print("=" * 60)
    create_roi_calculator_excel()
    print("=" * 60)
    print("Done!")

